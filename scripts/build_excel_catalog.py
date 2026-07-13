# -*- coding: utf-8 -*-
"""Convert the curated controls spreadsheet (Title + Description) into a
Raw_Catalogs envelope that ingest_raw.py understands, so the 761 controls enter
the pipeline as raw_artifacts with full lineage + content hashes (no bespoke raw
writer). Deterministic: row order defines the raw id (`<catalog>::NNNN`)."""
import argparse
import io
import json
import os
import re

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DEFAULT_XLSX = os.path.join(ROOT, 'cleaned_security_controls_table-ok.xlsx')
OUT = os.path.join(ROOT, 'SecureGuide_Mobile_Docs', 'Raw_Catalogs', 'securekit_curated_controls.json')
SOURCE_DOC = 'SecureGuide Curated Controls v1'

STOP = {'the', 'a', 'an', 'and', 'or', 'of', 'to', 'in', 'for', 'on', 'is', 'be', 'must',
        'that', 'this', 'with', 'as', 'at', 'by', 'are', 'all', 'any', 'from', 'which'}


def keywords(title, desc):
    words = re.findall(r"[A-Za-z][A-Za-z\-]{2,}", f"{title} {desc}".lower())
    out = []
    for w in words:
        if w not in STOP and w not in out:
            out.append(w)
        if len(out) >= 8:
            break
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default=DEFAULT_XLSX)
    ap.add_argument('--out', default=OUT)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True)][1:]  # skip header

    artifacts = []
    for i, r in enumerate(rows):
        title = (r[0] or '').strip()
        desc = (str(r[1]).strip() if r[1] is not None else '')
        if not title and not desc:
            continue
        artifacts.append({
            'raw_artifact_id': f"SKC-{i:04d}",
            'source_metadata': {
                'source_document': SOURCE_DOC, 'source_type': 'STANDARD',
                'source_version': '1', 'source_section': str(i + 1),
            },
            'original_content': {
                'raw_text_en': desc or title,
                'original_heading': title,
                'context_paragraph': None,
            },
            'extracted_elements': {
                'title_draft': title,
                'description_draft': desc,
                'keywords': keywords(title, desc),
            },
        })

    envelope = {
        'extraction_metadata': {'source_document': SOURCE_DOC, 'total_artifacts': len(artifacts)},
        'artifacts': artifacts,
    }
    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    io.open(args.out, 'w', encoding='utf-8', newline='\n').write(
        json.dumps(envelope, ensure_ascii=False, indent=2))
    print(f"wrote {len(artifacts)} curated controls -> {args.out}")


if __name__ == '__main__':
    main()
