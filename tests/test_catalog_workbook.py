"""Schema and export/validation tests for the comprehensive catalog workbook."""

from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from secureguide.catalog_workbook import (
    TABLE_SHEETS,
    annotate_validation_errors,
    export_workbook,
    validate_workbook,
)
from secureguide.database import apply_migrations


EXPECTED_SHEETS = [
    "00_Manifest", "01_Artifacts", "02_Source_Lineage",
    "03_Framework_Mappings", "04_Relationships", "05_Tags",
    "06_Type_Specific", "07_Reference_Lists", "08_Validation_Errors",
    "09_Raw_Dispositions", "10_Applicability", "11_Reference_Assessments",
    "12_Technical_Dependencies", "13_Verification_Tools", "14_Stakeholders",
    "15_Remediation_Actions", "16_External_References", "17_Localizations",
    "18_Actions", "19_Variants", "20_Security_Objectives", "21_CSF_Functions",
    "22_Control_Purposes", "23_Implementation_Types", "24_Maturity_Requirements",
    "25_Verification_Evidence", "26_Threats", "27_Platforms",
    "28_Legacy_Assets", "29_Legacy_Provenance", "30_Artifact_ID_Aliases",
    "31_Source_Catalogs", "32_Source_Manifests", "33_Source_Rights",
]

EXPECTED_DETAIL_TABLES = {
    "10_Applicability": "artifact_applicability_scope",
    "11_Reference_Assessments": "artifact_self_assessments",
    "12_Technical_Dependencies": "technical_dependencies",
    "13_Verification_Tools": "verification_tools",
    "14_Stakeholders": "stakeholders",
    "15_Remediation_Actions": "remediation_actions",
    "16_External_References": "external_references",
    "17_Localizations": "artifact_localizations",
    "18_Actions": "artifact_actions",
    "19_Variants": "artifact_variants",
    "20_Security_Objectives": "artifact_security_objectives",
    "21_CSF_Functions": "artifact_csf_functions",
    "22_Control_Purposes": "artifact_control_purposes",
    "23_Implementation_Types": "artifact_implementation_types",
    "24_Maturity_Requirements": "artifact_maturity_requirements",
    "25_Verification_Evidence": "artifact_verification_evidence_types",
    "26_Threats": "artifact_threats",
    "27_Platforms": "artifact_platforms",
    "28_Legacy_Assets": "catalog_legacy_assets",
    "29_Legacy_Provenance": "catalog_legacy_provenance",
    "30_Artifact_ID_Aliases": "catalog_artifact_id_aliases",
    "31_Source_Catalogs": "source_catalogs",
    "32_Source_Manifests": "source_import_manifests",
    "33_Source_Rights": "source_rights_versions",
}


class WorkbookSchemaTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.db = Path(self.temp.name) / "workbook.db"
        apply_migrations(self.db)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_audit_schema_is_additive_and_constrained(self) -> None:
        conn = sqlite3.connect(self.db)
        try:
            tables = {row[0] for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )}
            self.assertIn("catalog_workbook_runs", tables)
            self.assertIn("catalog_workbook_row_audit", tables)
            conn.execute(
                """INSERT INTO catalog_workbook_runs(
                       id,operation,workbook_path,baseline_db_sha256,status,actor
                   ) VALUES('RUN','EXPORT','catalog.xlsx',?,'STARTED','tester')""",
                ("a" * 64,),
            )
            with self.assertRaises(sqlite3.IntegrityError):
                conn.execute(
                    """INSERT INTO catalog_workbook_row_audit(
                           run_id,sheet_name,row_key,action,outcome
                       ) VALUES('RUN','01_Artifacts','A','DELETE','VALID')"""
                )
        finally:
            conn.close()

    def _seed_catalog(self) -> None:
        conn = sqlite3.connect(self.db)
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            conn.execute(
                "INSERT INTO source_catalogs(id,name,source_type,version) VALUES('SRC','Source','STANDARD','1')"
            )
            conn.execute(
                """INSERT INTO source_import_manifests(
                       id,source_catalog_id,source_version,source_file,source_sha256,
                       manifest_sha256,importer_name,importer_version,raw_record_count
                   ) VALUES('MAN','SRC','1','source.json',?,?, 'test','1',1)""",
                ("a" * 64, "b" * 64),
            )
            conn.execute(
                """INSERT INTO source_rights_versions(
                       id,source_catalog_id,source_version,rights_version,
                       redistribution_status,ship_raw_text,decision_reason,decided_by
                   ) VALUES('RIGHT','SRC','1','1','UNKNOWN',0,'No evidence','test')"""
            )
            conn.execute(
                """INSERT INTO raw_artifacts(
                       id,source_catalog_id,source_document,source_type,source_version,
                       source_section,raw_text_en,raw_json,source_file,content_hash,source_manifest_id
                   ) VALUES('RAW','SRC','Source','STANDARD','1','1','raw','{}','source.json',?,'MAN')""",
                ("c" * 64,),
            )
            conn.execute(
                """INSERT INTO security_artifacts(
                       id,source_catalog_id,type,title_en,definition_short_en,
                       primary_domain,sub_domain,abstraction_level,source,source_type,
                       obligation_level,granularity_level,control_nature,control_function,
                       testability,classification_confidence,classification_rationale,
                       ai_review_status,requires_human_review,publication_status,source_document
                   ) VALUES('SG-CTR-1','SRC','ART-CTR','Maintain inventory','Maintain an accurate inventory.',
                            'SD-02','SD-02.01','ABS-CTR','SRC-STD','STANDARD','OBL-MND',
                            'GRN-MEDIUM','NAT-ORG','FUN-PRE','TST-MAN',0.65,
                            'A safeguard that reduces inventory risk.','AIR-HUMAN-REVIEW',1,'APPROVED','Source')"""
            )
            conn.execute(
                """INSERT INTO artifact_source_lineage(
                       artifact_id,raw_artifact_id,lineage_role,mapping_strength,is_primary
                   ) VALUES('SG-CTR-1','RAW','SUPPORTS_CANONICAL','DIRECT',1)"""
            )
            conn.execute(
                """INSERT INTO raw_artifact_dispositions(
                       raw_artifact_id,disposition,rationale,decision_method,
                       decision_confidence,requires_human_review,decided_by
                   ) VALUES('RAW','SUPPORTS_CANONICAL','Primary source','TEST',0.65,1,'test')"""
            )
            conn.commit()
        finally:
            conn.close()

    def _seed_comprehensive_details(self) -> None:
        conn = sqlite3.connect(self.db)
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            conn.executescript(
                """
                INSERT INTO artifact_applicability_scope(artifact_id,scope_type,scope_value)
                VALUES('SG-CTR-1','INDUSTRY','Finance');
                INSERT INTO artifact_self_assessments(artifact_id,status,score,assessed_by)
                VALUES('SG-CTR-1','COMPLETED',90,'tester');
                INSERT INTO technical_dependencies(artifact_id,dependency_type,dependency_name,dependency_status)
                VALUES('SG-CTR-1','SYSTEM','Inventory service','AVAILABLE');
                INSERT INTO verification_tools(artifact_id,tool_name,tool_type,verification_method)
                VALUES('SG-CTR-1','SIEM','SIEM','LOG');
                INSERT INTO stakeholders(artifact_id,role,responsibility)
                VALUES('SG-CTR-1','Asset owner','OWNER');
                INSERT INTO remediation_actions(artifact_id,action,priority,effort_estimate,responsible_role)
                VALUES('SG-CTR-1','Update inventory','PRI-HIGH',2,'Asset owner');
                INSERT INTO external_references(artifact_id,type,title,url)
                VALUES('SG-CTR-1','ARTICLE','Inventory guide','https://example.invalid/inventory');
                INSERT INTO artifact_localizations(artifact_id,locale,is_primary,title)
                VALUES('SG-CTR-1','ar',0,'حصر الأصول');
                INSERT INTO artifact_variants(artifact_id,platform,title_en)
                VALUES('SG-CTR-1','generic','Generic inventory');
                INSERT INTO artifact_actions(artifact_id,kind,seq,text_en)
                VALUES('SG-CTR-1','ACTION',1,'Collect inventory records.');
                INSERT INTO artifact_security_objectives(artifact_id,objective_code,strength)
                VALUES('SG-CTR-1','accountability','primary');
                INSERT INTO artifact_csf_functions(artifact_id,csf_code,strength)
                VALUES('SG-CTR-1','identify','primary');
                INSERT INTO artifact_control_purposes(artifact_id,purpose_code)
                VALUES('SG-CTR-1','preventive');
                INSERT INTO artifact_implementation_types(artifact_id,impl_type_code)
                VALUES('SG-CTR-1','administrative');
                INSERT INTO artifact_maturity_requirements(artifact_id,tier_code,objective_en)
                VALUES('SG-CTR-1','essential','Maintain the inventory.');
                INSERT INTO artifact_verification_evidence_types(artifact_id,evidence_type)
                VALUES('SG-CTR-1','REPORT');
                INSERT INTO artifact_threats(artifact_id,threat_code)
                SELECT 'SG-CTR-1',code FROM lk_threat ORDER BY sort_order,code LIMIT 1;
                INSERT INTO artifact_platforms(artifact_id,platform_code)
                SELECT 'SG-CTR-1',code FROM lk_platform ORDER BY sort_order,code LIMIT 1;
                INSERT INTO catalog_legacy_assets(artifact_id,asset_ref)
                VALUES('SG-CTR-1','ASSET-001');
                INSERT INTO catalog_legacy_provenance(artifact_id,legacy_id,legacy_domain,legacy_sub)
                SELECT 'SG-CTR-1','LEGACY-001',legacy_key,'Sub' FROM legacy_domain_alias
                ORDER BY legacy_key LIMIT 1;
                """
            )
            conn.commit()
        finally:
            conn.close()

    def test_export_has_comprehensive_sheets_named_lists_and_validates(self) -> None:
        self._seed_catalog()
        self._seed_comprehensive_details()
        # openpyxl reloads an Excel numeric 0.0 as int 0; the semantic row
        # envelope must not report that representation change as an edit.
        conn = sqlite3.connect(self.db)
        conn.execute(
            "UPDATE security_artifacts SET classification_confidence=0.0 "
            "WHERE id='SG-CTR-1'"
        )
        conn.commit()
        conn.close()
        workbook = Path(self.temp.name) / "catalog.xlsx"
        result = export_workbook(self.db, workbook)
        self.assertEqual(result["sheets"], EXPECTED_SHEETS)
        wb = load_workbook(workbook, data_only=False)
        self.assertEqual(wb.sheetnames, EXPECTED_SHEETS)
        self.assertEqual(
            {sheet: TABLE_SHEETS[sheet] for sheet in EXPECTED_DETAIL_TABLES},
            EXPECTED_DETAIL_TABLES,
        )
        conn = sqlite3.connect(self.db)
        try:
            for sheet, table in TABLE_SHEETS.items():
                expected_columns = [
                    row[1] for row in conn.execute(f'PRAGMA table_info("{table}")')
                ]
                expected_rows = conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
                self.assertEqual(
                    [cell.value for cell in wb[sheet][1]],
                    ["_action", "_baseline_key", "_baseline_hash", *expected_columns],
                    sheet,
                )
                self.assertEqual(wb[sheet].max_row - 1, expected_rows, sheet)
        finally:
            conn.close()
        self.assertIn("REF_TYPE", wb.defined_names)
        self.assertIn("REF_EXTERNAL_REFERENCE_TYPE", wb.defined_names)
        self.assertGreater(len(wb["01_Artifacts"].data_validations.dataValidation), 1)
        self.assertGreater(len(wb["16_External_References"].data_validations.dataValidation), 0)
        manifest = {
            row[0].value: row[1].value
            for row in wb["00_Manifest"].iter_rows(min_row=2)
            if row[0].value
        }
        self.assertEqual(manifest["workbook_contract"], "secureguide-catalog-workbook-v3")
        self.assertTrue(str(manifest["database_path"]).endswith(self.db.name))
        self.assertEqual(manifest["artifact_count"], 1)
        self.assertEqual(manifest["row_count.17_Localizations"], 2)
        self.assertEqual(manifest["export_scope"], "ALL_CATALOG_ARTIFACTS")
        validation = validate_workbook(workbook, self.db)
        self.assertTrue(validation["valid"], validation["errors"][:3])

    def test_formula_and_edit_without_action_are_rejected(self) -> None:
        self._seed_catalog()
        workbook = Path(self.temp.name) / "catalog.xlsx"
        export_workbook(self.db, workbook)
        wb = load_workbook(workbook)
        ws = wb["01_Artifacts"]
        headers = {cell.value: cell.column for cell in ws[1]}
        ws.cell(2, headers["title_en"], "=\"Changed\"")
        wb.save(workbook)
        result = validate_workbook(workbook, self.db)
        codes = {error["code"] for error in result["errors"]}
        self.assertIn("FORMULA", codes)
        self.assertIn("ACTION_REQUIRED", codes)
        annotated = Path(self.temp.name) / "catalog.validated.xlsx"
        annotation = annotate_validation_errors(workbook, result["errors"], annotated)
        self.assertEqual(annotation["errorCount"], len(result["errors"]))
        annotated_wb = load_workbook(annotated, data_only=False)
        error_rows = list(
            annotated_wb["08_Validation_Errors"].iter_rows(
                min_row=2, values_only=True
            )
        )
        self.assertEqual(len(error_rows), len(result["errors"]))
        self.assertEqual(
            {row[3] for row in error_rows},
            {error["code"] for error in result["errors"]},
        )

    def test_empty_string_detail_round_trips_as_untouched_blank_cell(self) -> None:
        """XLSX reloads a stored empty string as None, without an edit."""
        self._seed_catalog()
        conn = sqlite3.connect(self.db)
        try:
            conn.execute(
                """INSERT INTO framework_mappings(
                       artifact_id,framework,version,reference,mapping_strength,rationale
                   ) VALUES('SG-CTR-1','SecureGuide Curated Controls v1','1','1','DIRECT','')"""
            )
            conn.commit()
        finally:
            conn.close()
        workbook = Path(self.temp.name) / "empty-string-detail.xlsx"
        export_workbook(self.db, workbook)
        result = validate_workbook(workbook, self.db)
        self.assertTrue(result["valid"], result["errors"])

    def test_filtered_export_is_lossless_for_artifact_and_raw_scope(self) -> None:
        self._seed_catalog()
        conn = sqlite3.connect(self.db)
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            conn.execute(
                """INSERT INTO raw_artifacts(
                       id,source_catalog_id,source_document,source_type,source_version,
                       source_section,raw_text_en,raw_json,source_file,content_hash,source_manifest_id)
                   VALUES('RAW-2','SRC','Source','STANDARD','1','2','raw two','{}','source.json',?,'MAN')""",
                ("d" * 64,),
            )
            conn.execute(
                """INSERT INTO security_artifacts(
                       id,source_catalog_id,type,title_en,definition_short_en,
                       primary_domain,sub_domain,abstraction_level,source,source_type,
                       obligation_level,granularity_level,requirement_type,
                       classification_confidence,classification_rationale,
                       ai_review_status,requires_human_review,publication_status,source_document)
                   VALUES('SG-REQ-2','SRC','ART-REQ','Govern access','Access shall be governed.',
                          'SD-03','SD-03.01','ABS-GOV','SRC-STD','STANDARD','OBL-MND',
                          'GRN-HIGH','RQT-STD',.91,'Security outcome.','AIR-AUTO-ACCEPTED',0,
                          'APPROVED','Source')"""
            )
            conn.execute(
                "INSERT INTO artifact_source_lineage(artifact_id,raw_artifact_id,lineage_role,mapping_strength,is_primary) VALUES('SG-REQ-2','RAW-2','SUPPORTS_CANONICAL','DIRECT',1)"
            )
            conn.execute(
                "INSERT INTO raw_artifact_dispositions(raw_artifact_id,disposition,rationale,decision_method,decision_confidence,requires_human_review,decided_by) VALUES('RAW-2','SUPPORTS_CANONICAL','Primary source','TEST',.91,0,'test')"
            )
            conn.commit()
        finally:
            conn.close()

        workbook = Path(self.temp.name) / "requirements.xlsx"
        result = export_workbook(
            self.db, workbook, filters={"artifact_type": "ART-REQ"}
        )
        self.assertEqual(result["exportMode"], "FILTERED")
        wb = load_workbook(workbook, data_only=True)
        self.assertEqual(wb["01_Artifacts"].max_row - 1, 1)
        self.assertEqual(wb["02_Source_Lineage"].max_row - 1, 1)
        self.assertEqual(wb["09_Raw_Dispositions"].max_row - 1, 1)
        artifact_headers = [cell.value for cell in wb["01_Artifacts"][1]]
        raw_headers = [cell.value for cell in wb["09_Raw_Dispositions"][1]]
        self.assertEqual(
            wb["01_Artifacts"].cell(2, artifact_headers.index("id") + 1).value,
            "SG-REQ-2",
        )
        self.assertEqual(
            wb["09_Raw_Dispositions"].cell(
                2, raw_headers.index("raw_artifact_id") + 1
            ).value,
            "RAW-2",
        )
        validation = validate_workbook(workbook, self.db)
        self.assertTrue(validation["valid"], validation["errors"][:3])

    def test_duplicate_editable_row_identity_is_rejected(self) -> None:
        self._seed_catalog()
        workbook = Path(self.temp.name) / "catalog.xlsx"
        export_workbook(self.db, workbook)
        wb = load_workbook(workbook)
        ws = wb["01_Artifacts"]
        ws.append([cell.value for cell in ws[2]])
        wb.save(workbook)
        result = validate_workbook(workbook, self.db)
        duplicate = [
            error
            for error in result["errors"]
            if error["code"] == "DUPLICATE_ROW_KEY"
        ]
        self.assertEqual(len(duplicate), 1)
        self.assertEqual(duplicate[0]["sheet"], "01_Artifacts")
        self.assertIn("row 2", duplicate[0]["message"])


if __name__ == "__main__":
    unittest.main()
