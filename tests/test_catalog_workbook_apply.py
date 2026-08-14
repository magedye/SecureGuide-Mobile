"""Conflict, omission, rollback, and preservation tests for workbook apply."""

from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from secureguide.catalog_workbook import (
    WorkbookConflict,
    WorkbookError,
    apply_workbook_plan,
    catalog_state_hash,
    export_workbook,
    plan_workbook,
)
from secureguide.database import apply_migrations, connect


def seed(database: Path) -> None:
    conn = sqlite3.connect(database)
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        INSERT INTO source_catalogs(id,name,source_type,version)
        VALUES('SRC','Source','STANDARD','1');
        INSERT INTO source_import_manifests(
          id,source_catalog_id,source_version,source_file,source_sha256,
          manifest_sha256,importer_name,importer_version,raw_record_count)
        VALUES('MAN','SRC','1','source.json',
          'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa',
          'bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb',
          'test','1',1);
        INSERT INTO source_rights_versions(
          id,source_catalog_id,source_version,rights_version,
          redistribution_status,ship_raw_text,decision_reason,decided_by)
        VALUES('RIGHT','SRC','1','1','UNKNOWN',0,'No evidence','test');
        INSERT INTO raw_artifacts(
          id,source_catalog_id,source_document,source_type,source_version,
          source_section,raw_text_en,raw_json,source_file,content_hash,source_manifest_id)
        VALUES('RAW','SRC','Source','STANDARD','1','1','raw','{}','source.json',
          'cccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccccc','MAN');
        INSERT INTO security_artifacts(
          id,source_catalog_id,type,title_en,definition_short_en,
          primary_domain,sub_domain,abstraction_level,source,source_type,
          obligation_level,granularity_level,control_nature,control_function,
          testability,classification_confidence,classification_rationale,
          ai_review_status,requires_human_review,publication_status,source_document)
        VALUES('SG-CTR-1','SRC','ART-CTR','Maintain inventory','Maintain an accurate inventory.',
          'SD-02','SD-02.01','ABS-CTR','SRC-STD','STANDARD','OBL-MND',
          'GRN-MEDIUM','NAT-ORG','FUN-PRE','TST-MAN',0.65,
          'A safeguard that reduces inventory risk.','AIR-HUMAN-REVIEW',1,'APPROVED','Source');
        INSERT INTO artifact_source_lineage(
          artifact_id,raw_artifact_id,lineage_role,mapping_strength,is_primary)
        VALUES('SG-CTR-1','RAW','SUPPORTS_CANONICAL','DIRECT',1);
        INSERT INTO raw_artifact_dispositions(
          raw_artifact_id,disposition,rationale,decision_method,
          decision_confidence,requires_human_review,decided_by)
        VALUES('RAW','SUPPORTS_CANONICAL','Primary source','TEST',0.65,1,'test');
        """
    )
    conn.commit()
    conn.close()


def edit_artifact(workbook: Path, field: str, value: object, action: str = "UPSERT") -> None:
    wb = load_workbook(workbook)
    ws = wb["01_Artifacts"]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["_action"], action)
    ws.cell(2, headers[field], value)
    wb.save(workbook)


def edit_detail(
    workbook: Path, sheet: str, field: str, value: object, action: str = "UPSERT"
) -> None:
    wb = load_workbook(workbook)
    ws = wb[sheet]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["_action"], action)
    ws.cell(2, headers[field], value)
    wb.save(workbook)


class WorkbookApplyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.db = Path(self.temp.name) / "workbook.db"
        self.workbook = Path(self.temp.name) / "catalog.xlsx"
        apply_migrations(self.db)
        seed(self.db)

    def tearDown(self) -> None:
        self.temp.cleanup()

    def test_valid_edit_round_trip_is_audited(self) -> None:
        export_workbook(self.db, self.workbook)
        edit_artifact(self.workbook, "title_en", "Updated inventory title")
        plan = plan_workbook(self.workbook, self.db)
        self.assertEqual(plan["conflicts"], [])
        result = apply_workbook_plan(plan, actor="tester")
        conn = sqlite3.connect(self.db)
        try:
            self.assertEqual(result["status"], "APPLIED")
            self.assertEqual(conn.execute(
                "SELECT title_en FROM security_artifacts WHERE id='SG-CTR-1'"
            ).fetchone()[0], "Updated inventory title")
            self.assertEqual(conn.execute(
                "SELECT COUNT(*) FROM catalog_workbook_row_audit WHERE outcome='APPLIED'"
            ).fetchone()[0], 1)
        finally:
            conn.close()

    def test_raw_disposition_round_trip_is_transactional_and_audited(self) -> None:
        export_workbook(self.db, self.workbook)
        edit_detail(
            self.workbook,
            "09_Raw_Dispositions",
            "rationale",
            "Reviewed primary source disposition.",
        )
        plan = plan_workbook(self.workbook, self.db)
        result = apply_workbook_plan(plan, actor="reviewer")
        conn = sqlite3.connect(self.db)
        try:
            self.assertEqual(result["status"], "APPLIED")
            self.assertEqual(
                conn.execute(
                    "SELECT rationale FROM raw_artifact_dispositions WHERE raw_artifact_id='RAW'"
                ).fetchone()[0],
                "Reviewed primary source disposition.",
            )
            self.assertEqual(
                conn.execute(
                    "SELECT sheet_name FROM catalog_workbook_row_audit"
                ).fetchone()[0],
                "09_Raw_Dispositions",
            )
        finally:
            conn.close()

    def test_comprehensive_detail_edit_round_trip_is_audited(self) -> None:
        conn = sqlite3.connect(self.db)
        conn.execute(
            """INSERT INTO technical_dependencies(
                   artifact_id,dependency_type,dependency_name,dependency_status)
               VALUES('SG-CTR-1','SYSTEM','Inventory service','AVAILABLE')"""
        )
        conn.commit()
        conn.close()
        export_workbook(self.db, self.workbook)
        edit_detail(
            self.workbook,
            "12_Technical_Dependencies",
            "dependency_status",
            "PLANNED",
        )
        plan = plan_workbook(self.workbook, self.db)
        self.assertEqual(plan["contract"], "secureguide-catalog-workbook-plan-v3")
        self.assertEqual(plan["conflicts"], [])
        result = apply_workbook_plan(plan, actor="tester")
        conn = sqlite3.connect(self.db)
        try:
            self.assertEqual(result["affectedArtifacts"], 1)
            self.assertEqual(
                conn.execute(
                    "SELECT dependency_status FROM technical_dependencies "
                    "WHERE artifact_id='SG-CTR-1'"
                ).fetchone()[0],
                "PLANNED",
            )
            self.assertEqual(
                conn.execute(
                    "SELECT sheet_name FROM catalog_workbook_row_audit"
                ).fetchone()[0],
                "12_Technical_Dependencies",
            )
        finally:
            conn.close()

    def test_child_detail_cannot_be_deprecated(self) -> None:
        conn = sqlite3.connect(self.db)
        conn.execute(
            """INSERT INTO technical_dependencies(
                   artifact_id,dependency_type,dependency_name,dependency_status)
               VALUES('SG-CTR-1','SYSTEM','Inventory service','AVAILABLE')"""
        )
        conn.commit()
        conn.close()
        export_workbook(self.db, self.workbook)
        edit_detail(
            self.workbook,
            "12_Technical_Dependencies",
            "dependency_status",
            "AVAILABLE",
            action="DEPRECATE",
        )
        with self.assertRaisesRegex(WorkbookError, "validation failed"):
            plan_workbook(self.workbook, self.db)

    def test_row_omission_is_no_change(self) -> None:
        export_workbook(self.db, self.workbook)
        wb = load_workbook(self.workbook)
        wb["01_Artifacts"].delete_rows(2)
        wb.save(self.workbook)
        conn = connect(self.db)
        before = catalog_state_hash(conn)
        conn.close()
        result = apply_workbook_plan(plan_workbook(self.workbook, self.db))
        conn = connect(self.db)
        after = catalog_state_hash(conn)
        conn.close()
        self.assertEqual(result["entries"], 0)
        self.assertEqual(before, after)

    def test_stale_database_requires_explicit_resolution(self) -> None:
        export_workbook(self.db, self.workbook)
        edit_artifact(self.workbook, "title_en", "Workbook title")
        conn = sqlite3.connect(self.db)
        conn.execute("UPDATE security_artifacts SET title_en='Database title' WHERE id='SG-CTR-1'")
        conn.commit()
        conn.close()
        unresolved = plan_workbook(self.workbook, self.db)
        self.assertEqual(len(unresolved["conflicts"]), 1)
        with self.assertRaises(WorkbookConflict):
            apply_workbook_plan(unresolved)
        resolved = plan_workbook(self.workbook, self.db, {"__database__": "USE_DATABASE"})
        apply_workbook_plan(resolved)
        conn = sqlite3.connect(self.db)
        self.assertEqual(conn.execute(
            "SELECT title_en FROM security_artifacts WHERE id='SG-CTR-1'"
        ).fetchone()[0], "Database title")
        conn.close()

    def test_mid_batch_failure_rolls_back_first_edit(self) -> None:
        export_workbook(self.db, self.workbook)
        edit_artifact(self.workbook, "title_en", "Must roll back")
        wb = load_workbook(self.workbook)
        ws = wb["02_Source_Lineage"]
        headers = {cell.value: cell.column for cell in ws[1]}
        row = ws.max_row + 1
        ws.cell(row, headers["_action"], "UPSERT")
        ws.cell(row, headers["artifact_id"], "MISSING")
        ws.cell(row, headers["raw_artifact_id"], "RAW")
        ws.cell(row, headers["lineage_role"], "SUPPORTS_CANONICAL")
        ws.cell(row, headers["mapping_strength"], "DIRECT")
        ws.cell(row, headers["is_primary"], 0)
        wb.save(self.workbook)
        plan = plan_workbook(self.workbook, self.db)
        with self.assertRaises(sqlite3.IntegrityError):
            apply_workbook_plan(plan)
        conn = sqlite3.connect(self.db)
        self.assertEqual(conn.execute(
            "SELECT title_en FROM security_artifacts WHERE id='SG-CTR-1'"
        ).fetchone()[0], "Maintain inventory")
        self.assertEqual(conn.execute("SELECT COUNT(*) FROM catalog_workbook_runs").fetchone()[0], 0)
        conn.close()

    def test_deprecation_preserves_profile_reference(self) -> None:
        conn = sqlite3.connect(self.db)
        try:
            conn.execute("INSERT INTO enterprise_profiles(id,name,profile_kind) VALUES('P','Profile','organization')")
            conn.execute("INSERT INTO profile_artifacts(id,profile_id,artifact_id) VALUES('PA','P','SG-CTR-1')")
            conn.commit()
        finally:
            conn.close()
        export_workbook(self.db, self.workbook)
        edit_artifact(self.workbook, "title_en", "Maintain inventory", action="DEPRECATE")
        apply_workbook_plan(plan_workbook(self.workbook, self.db))
        conn = sqlite3.connect(self.db)
        self.assertEqual(conn.execute(
            "SELECT publication_status,is_active FROM security_artifacts WHERE id='SG-CTR-1'"
        ).fetchone(), ("DEPRECATED", 0))
        self.assertEqual(conn.execute("SELECT COUNT(*) FROM profile_artifacts").fetchone()[0], 1)
        conn.close()

    def test_release_asset_is_never_a_direct_apply_target(self) -> None:
        with self.assertRaisesRegex(WorkbookError, "mobile/assets"):
            apply_workbook_plan({"database": str(Path("mobile/assets/catalog.db").resolve())})


if __name__ == "__main__":
    unittest.main()
