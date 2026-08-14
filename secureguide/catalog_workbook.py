"""Audited comprehensive Excel round-trip for SecureGuide catalog curation."""

from __future__ import annotations

import json
import sqlite3
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from secureguide.catalog_validation import (
    canonical_hash,
    file_hash,
    load_contract,
    minimum_result,
)
from secureguide.database import connect


ROOT = Path(__file__).resolve().parent.parent
RELEASE_ASSET = (ROOT / "mobile" / "assets" / "catalog.db").resolve()
CORE_SHEETS = (
    "00_Manifest", "01_Artifacts", "02_Source_Lineage",
    "03_Framework_Mappings", "04_Relationships", "05_Tags",
    "06_Type_Specific", "07_Reference_Lists", "08_Validation_Errors",
    "09_Raw_Dispositions",
)
DETAIL_TABLE_SHEETS = {
    "10_Applicability": "artifact_applicability_scope",
    "11_Reference_Assessments": "artifact_self_assessments",
    "12_Technical_Dependencies": "technical_dependencies",
    "13_Verification_Tools": "verification_tools",
    "14_Stakeholders": "stakeholders",
    "15_Remediation_Actions": "remediation_actions",
    "16_External_References": "external_references",
    "17_Localizations": "artifact_localizations",
    "18_Actions": "artifact_actions",
    "19_Variants": "artifact_variants",
    "20_Security_Objectives": "artifact_security_objectives",
    "21_CSF_Functions": "artifact_csf_functions",
    "22_Control_Purposes": "artifact_control_purposes",
    "23_Implementation_Types": "artifact_implementation_types",
    "24_Maturity_Requirements": "artifact_maturity_requirements",
    "25_Verification_Evidence": "artifact_verification_evidence_types",
    "26_Threats": "artifact_threats",
    "27_Platforms": "artifact_platforms",
    "28_Legacy_Assets": "catalog_legacy_assets",
    "29_Legacy_Provenance": "catalog_legacy_provenance",
    "30_Artifact_ID_Aliases": "catalog_artifact_id_aliases",
    "31_Source_Catalogs": "source_catalogs",
    "32_Source_Manifests": "source_import_manifests",
    "33_Source_Rights": "source_rights_versions",
}
SHEETS = (*CORE_SHEETS, *DETAIL_TABLE_SHEETS)
ACTION_VALUES = ("NO_CHANGE", "UPSERT", "DEPRECATE")
TABLE_SHEETS = {
    "01_Artifacts": "security_artifacts",
    "02_Source_Lineage": "artifact_source_lineage",
    "03_Framework_Mappings": "framework_mappings",
    "04_Relationships": "artifact_relationships",
    "05_Tags": "artifact_tags",
    "09_Raw_Dispositions": "raw_artifact_dispositions",
    **DETAIL_TABLE_SHEETS,
}
EDITABLE_SHEETS = (
    "01_Artifacts", "02_Source_Lineage", "03_Framework_Mappings",
    "04_Relationships", "05_Tags", "06_Type_Specific",
    "09_Raw_Dispositions",
    *tuple(sheet for sheet in DETAIL_TABLE_SHEETS if not sheet.startswith(("30_", "31_", "32_", "33_"))),
)
PRIMARY_KEYS = {
    "01_Artifacts": ("id",),
    "02_Source_Lineage": ("artifact_id", "raw_artifact_id"),
    "03_Framework_Mappings": ("id",),
    "04_Relationships": ("id",),
    "05_Tags": ("artifact_id", "tag_type", "tag_value"),
    "06_Type_Specific": ("artifact_id",),
    "09_Raw_Dispositions": ("raw_artifact_id",),
    "10_Applicability": ("id",),
    "11_Reference_Assessments": ("id",),
    "12_Technical_Dependencies": ("id",),
    "13_Verification_Tools": ("id",),
    "14_Stakeholders": ("id",),
    "15_Remediation_Actions": ("id",),
    "16_External_References": ("id",),
    "17_Localizations": ("artifact_id", "locale"),
    "18_Actions": ("id",),
    "19_Variants": ("id",),
    "20_Security_Objectives": ("artifact_id", "objective_code"),
    "21_CSF_Functions": ("artifact_id", "csf_code"),
    "22_Control_Purposes": ("artifact_id", "purpose_code"),
    "23_Implementation_Types": ("artifact_id", "impl_type_code"),
    "24_Maturity_Requirements": ("id",),
    "25_Verification_Evidence": ("artifact_id", "evidence_type"),
    "26_Threats": ("artifact_id", "threat_code"),
    "27_Platforms": ("artifact_id", "platform_code"),
    "28_Legacy_Assets": ("artifact_id", "asset_ref"),
    "29_Legacy_Provenance": ("artifact_id",),
    "30_Artifact_ID_Aliases": ("old_artifact_id",),
    "31_Source_Catalogs": ("id",),
    "32_Source_Manifests": ("id",),
    "33_Source_Rights": ("id",),
}
TYPE_FIELDS = (
    "artifact_id", "type", "requirement_type", "control_nature",
    "control_function", "testability", "asset_type", "asset_criticality",
    "exception_approval_date", "exception_expiry_date", "effective_date",
)
LOOKUPS = {
    "type": "lk_artifact_type", "primary_domain": "lk_sdt_domain",
    "sub_domain": "lk_sdt_subdomain", "abstraction_level": "lk_abstraction_level",
    "source": "lk_obligation_source", "source_type": "lk_source_type",
    "obligation_level": "lk_obligation_level", "requirement_type": "lk_requirement_type",
    "granularity_level": "lk_granularity_level", "control_nature": "lk_control_nature",
    "control_function": "lk_control_function", "testability": "lk_testability",
    "priority": "lk_priority", "review_frequency": "lk_review_frequency",
    "publication_status": "lk_publication_status", "ai_review_status": "lk_ai_review_status",
    "implementation_status": "lk_implementation_status",
    "verification_status": "lk_verification_status",
    "effectiveness": "lk_effectiveness", "exception_status": "lk_exception_status",
    "asset_type": "lk_asset_type", "maturity_level": "lk_maturity_level",
    "cost_category": "lk_cost_category", "import_status": "lk_import_status",
    "resolution_status": "lk_resolution_status",
    "applicability_scope_type": "lk_applicability_scope_type",
    "self_assessment_status": "lk_self_assessment_status",
    "dependency_type": "lk_dependency_type", "dependency_status": "lk_dependency_status",
    "verification_tool_type": "lk_verification_tool_type",
    "verification_method": "lk_verification_method",
    "stakeholder_responsibility": "lk_stakeholder_responsibility",
    "external_reference_type": "lk_external_reference_type",
    "objective_code": "lk_security_objective",
    "objective_strength": "lk_objective_strength",
    "csf_code": "lk_csf_function", "purpose_code": "lk_control_purpose",
    "impl_type_code": "lk_implementation_type", "tier_code": "lk_tier",
    "threat_code": "lk_threat", "platform_code": "lk_platform",
}
STATIC_LISTS = {
    "action": ACTION_VALUES,
    "mapping_strength": ("DIRECT", "INDIRECT", "PARTIAL", "INFORMATIVE"),
    "lineage_role": ("SUPPORTS_CANONICAL", "SPLIT"),
    "relation_type": (
        "REL-DER", "REL-SAT", "REL-SUP", "REL-SPL", "REL-IMP", "REL-VER",
        "REL-MEA", "REL-MIT", "REL-AFF", "REL-EXC", "REL-DEP", "REL-CNF",
    ),
    "tag_type": ("Technology", "Framework", "Concept", "Context", "Threat", "Data", "Party"),
    "content_maturity": ("DRAFT", "MINIMAL", "ENRICHED", "REVIEWED"),
    "content_review_status": ("NOT_REVIEWED", "NEEDS_REVIEW", "APPROVED"),
    "action_kind": ("ACTION", "VERIFICATION"),
    "csf_strength": ("primary", "supporting"),
    "verification_evidence_type": (
        "DOCUMENT", "SCREENSHOT", "LOG", "REPORT", "CONFIG",
        "ATTESTATION", "LINK", "OTHER",
    ),
    "raw_disposition": tuple(load_contract()["raw_dispositions"]),
}
CUSTOM_LIST_QUERIES = {
    "legacy_domain": "SELECT legacy_key FROM legacy_domain_alias ORDER BY legacy_key",
}
GLOBAL_FIELD_LISTS = {
    "type": "type", "primary_domain": "primary_domain", "sub_domain": "sub_domain",
    "abstraction_level": "abstraction_level", "source": "source",
    "source_type": "source_type", "obligation_level": "obligation_level",
    "requirement_type": "requirement_type", "granularity_level": "granularity_level",
    "control_nature": "control_nature", "control_function": "control_function",
    "testability": "testability", "priority": "priority",
    "implementation_status": "implementation_status",
    "verification_status": "verification_status", "effectiveness": "effectiveness",
    "exception_status": "exception_status", "review_frequency": "review_frequency",
    "publication_status": "publication_status", "asset_type": "asset_type",
    "required_maturity_level": "maturity_level", "cost_category": "cost_category",
    "ai_review_status": "ai_review_status", "import_status": "import_status",
    "mapping_strength": "mapping_strength", "tag_type": "tag_type",
}
SHEET_FIELD_LISTS = {
    "02_Source_Lineage": {"lineage_role": "lineage_role"},
    "04_Relationships": {
        "relation_type": "relation_type", "resolution_status": "resolution_status",
    },
    "09_Raw_Dispositions": {"disposition": "raw_disposition"},
    "10_Applicability": {"scope_type": "applicability_scope_type"},
    "11_Reference_Assessments": {"status": "self_assessment_status"},
    "12_Technical_Dependencies": {
        "dependency_type": "dependency_type", "dependency_status": "dependency_status",
    },
    "13_Verification_Tools": {
        "tool_type": "verification_tool_type", "verification_method": "verification_method",
    },
    "14_Stakeholders": {"responsibility": "stakeholder_responsibility"},
    "16_External_References": {"type": "external_reference_type"},
    "17_Localizations": {
        "content_maturity": "content_maturity",
        "content_review_status": "content_review_status",
    },
    "18_Actions": {"kind": "action_kind"},
    "20_Security_Objectives": {
        "objective_code": "objective_code", "strength": "objective_strength",
    },
    "21_CSF_Functions": {"csf_code": "csf_code", "strength": "csf_strength"},
    "22_Control_Purposes": {"purpose_code": "purpose_code"},
    "23_Implementation_Types": {"impl_type_code": "impl_type_code"},
    "24_Maturity_Requirements": {"tier_code": "tier_code"},
    "25_Verification_Evidence": {"evidence_type": "verification_evidence_type"},
    "26_Threats": {"threat_code": "threat_code"},
    "27_Platforms": {"platform_code": "platform_code"},
    "29_Legacy_Provenance": {"legacy_domain": "legacy_domain"},
}
META_COLUMNS = ("_action", "_baseline_key", "_baseline_hash")


class WorkbookError(ValueError):
    pass


class WorkbookConflict(WorkbookError):
    pass


def _portable_path(path: str | Path) -> str:
    resolved = Path(path).resolve()
    try:
        return resolved.relative_to(ROOT.resolve()).as_posix()
    except ValueError:
        return str(resolved)


def _json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _row_dict(row: sqlite3.Row, columns: Iterable[str] | None = None) -> dict[str, Any]:
    names = columns or row.keys()
    return {name: _json_value(row[name]) for name in names}


def row_hash(values: dict[str, Any]) -> str:
    # Excel stores integral floating-point values such as 0.0 as numeric 0
    # and serializes an empty-string cell as a blank cell (None).  Treat both
    # representations as semantically equal so an untouched export can always
    # pass the no-op validation gate.  A deliberate UPSERT of a blank value
    # still carries the workbook's None to the database layer.
    normalized = {
        key: (
            int(values[key])
            if isinstance(values[key], float) and values[key].is_integer()
            else None if values[key] == "" else values[key]
        )
        for key in sorted(values)
    }
    return canonical_hash(normalized)


def _table_columns(conn: sqlite3.Connection, table: str) -> list[str]:
    columns = [row[1] for row in conn.execute(f'PRAGMA table_info("{table}")')]
    if not columns:
        raise WorkbookError(f"required workbook table is missing: {table}")
    return columns


def _table_rows(conn: sqlite3.Connection, table: str) -> list[dict[str, Any]]:
    columns = _table_columns(conn, table)
    order = ",".join(
        row[1] for row in conn.execute(f"PRAGMA table_info({table})") if row[5]
    ) or ",".join(columns)
    return [
        _row_dict(row, columns)
        for row in conn.execute(f'SELECT * FROM "{table}" ORDER BY {order}')
    ]


def _selected_artifacts(
    conn: sqlite3.Connection, filters: dict[str, Any]
) -> list[dict[str, Any]]:
    clauses: list[str] = []
    parameters: list[Any] = []
    exact_columns = {
        "artifact_type": "a.type", "primary_domain": "a.primary_domain",
        "sub_domain": "a.sub_domain", "publication_status": "a.publication_status",
        "ai_review_status": "a.ai_review_status",
    }
    for key, column in exact_columns.items():
        if filters.get(key) not in (None, ""):
            clauses.append(f"{column}=?")
            parameters.append(filters[key])
    if filters.get("requires_human_review") is not None:
        clauses.append("a.requires_human_review=?")
        parameters.append(int(bool(filters["requires_human_review"])))
    if filters.get("min_confidence") is not None:
        clauses.append("a.classification_confidence>=?")
        parameters.append(float(filters["min_confidence"]))
    if filters.get("max_confidence") is not None:
        clauses.append("a.classification_confidence<=?")
        parameters.append(float(filters["max_confidence"]))
    if filters.get("source") not in (None, ""):
        clauses.append(
            "(lower(a.source_catalog_id)=lower(?) OR lower(a.source)=lower(?) "
            "OR lower(a.source_document)=lower(?) OR EXISTS("
            "SELECT 1 FROM framework_mappings fm WHERE fm.artifact_id=a.id "
            "AND lower(fm.framework)=lower(?)))"
        )
        parameters.extend([filters["source"]] * 4)
    where = " WHERE " + " AND ".join(clauses) if clauses else ""
    columns = _table_columns(conn, "security_artifacts")
    rows = [
        _row_dict(row, columns)
        for row in conn.execute(
            f"SELECT a.* FROM security_artifacts a{where} ORDER BY a.id", parameters
        )
    ]
    quality = filters.get("quality_profile")
    if quality:
        from secureguide.catalog_validation import enriched_result, strict_result

        contract = load_contract()
        filtered: list[dict[str, Any]] = []
        for values in rows:
            row = conn.execute(
                "SELECT * FROM security_artifacts WHERE id=?", (values["id"],)
            ).fetchone()
            valid = {
                "MINIMUM_VALID": minimum_result(conn, row, contract)["valid"],
                "STRICT_USACM": strict_result(conn, row)["valid"],
                "ENRICHED": enriched_result(conn, row, contract)["valid"],
            }.get(str(quality))
            if valid is None:
                raise WorkbookError(f"unsupported quality profile filter: {quality}")
            if valid:
                filtered.append(values)
        rows = filtered
    return rows


def _rows_for_ids(
    conn: sqlite3.Connection,
    table: str,
    column: str,
    identifiers: set[str],
) -> list[dict[str, Any]]:
    if not identifiers:
        return []
    columns = _table_columns(conn, table)
    order = ",".join(
        row[1] for row in conn.execute(f"PRAGMA table_info({table})") if row[5]
    ) or ",".join(columns)
    placeholders = ",".join("?" for _ in identifiers)
    return [
        _row_dict(row, columns)
        for row in conn.execute(
            f'SELECT * FROM "{table}" WHERE "{column}" IN ({placeholders}) ORDER BY {order}',
            tuple(sorted(identifiers)),
        )
    ]


def _scoped_table_rows(
    conn: sqlite3.Connection,
    artifacts: list[dict[str, Any]],
    *,
    complete: bool,
) -> dict[str, list[dict[str, Any]]]:
    if complete:
        return {sheet: _table_rows(conn, table) for sheet, table in TABLE_SHEETS.items()}
    artifact_ids = {str(row["id"]) for row in artifacts}
    source_ids = {
        str(row["source_catalog_id"])
        for row in artifacts
        if row.get("source_catalog_id") not in (None, "")
    }
    result: dict[str, list[dict[str, Any]]] = {"01_Artifacts": artifacts}
    lineage = _rows_for_ids(
        conn, "artifact_source_lineage", "artifact_id", artifact_ids
    )
    raw_ids = {str(row["raw_artifact_id"]) for row in lineage}
    result["02_Source_Lineage"] = lineage
    for sheet, table in TABLE_SHEETS.items():
        if sheet in result:
            continue
        if sheet == "09_Raw_Dispositions":
            result[sheet] = _rows_for_ids(
                conn, table, "raw_artifact_id", raw_ids
            )
        elif sheet == "04_Relationships":
            source_rows = _rows_for_ids(conn, table, "source_id", artifact_ids)
            target_rows = _rows_for_ids(conn, table, "target_id", artifact_ids)
            by_id = {str(row["id"]): row for row in (*source_rows, *target_rows)}
            result[sheet] = [by_id[key] for key in sorted(by_id)]
        elif sheet in {"31_Source_Catalogs", "32_Source_Manifests", "33_Source_Rights"}:
            column = "id" if sheet == "31_Source_Catalogs" else "source_catalog_id"
            result[sheet] = _rows_for_ids(conn, table, column, source_ids)
        else:
            columns = _table_columns(conn, table)
            if "artifact_id" not in columns:
                raise WorkbookError(f"no scoped artifact key for {sheet}")
            result[sheet] = _rows_for_ids(conn, table, "artifact_id", artifact_ids)
    return result


def catalog_state_hash(conn: sqlite3.Connection) -> str:
    """Logical hash excluding workbook audit rows and SQLite byte layout."""
    payload = {table: _table_rows(conn, table) for table in TABLE_SHEETS.values()}
    return canonical_hash(payload)


def _style_sheet(ws, *, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    if ws.max_row and ws.max_column:
        ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    for column in range(1, ws.max_column + 1):
        header = str(ws.cell(1, column).value or "")
        width = min(52, max(12, len(header) + 2))
        for row in range(2, min(ws.max_row, 80) + 1):
            width = min(52, max(width, len(str(ws.cell(row, column).value or "")) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width


def _append_table(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append([*META_COLUMNS, *columns])
    for values in rows:
        key_columns = PRIMARY_KEYS[ws.title]
        baseline_key = canonical_hash({key: values.get(key) for key in key_columns})
        ws.append(["NO_CHANGE", baseline_key, row_hash(values), *[values.get(c) for c in columns]])


def _reference_lists(conn: sqlite3.Connection) -> dict[str, tuple[str, ...]]:
    tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    result = dict(STATIC_LISTS)
    for field, table in LOOKUPS.items():
        if table in tables:
            result[field] = tuple(row[0] for row in conn.execute(f"SELECT code FROM {table} ORDER BY sort_order,code"))
    for name, query in CUSTOM_LIST_QUERIES.items():
        result[name] = tuple(row[0] for row in conn.execute(query))
    return result


def _field_lists_for_sheet(sheet: str, columns: Iterable[str]) -> dict[str, str]:
    """Map worksheet columns to controlled-list names without name collisions."""
    overridden = SHEET_FIELD_LISTS.get(sheet, {})
    result: dict[str, str] = {}
    for field in columns:
        if field in overridden:
            result[field] = overridden[field]
        elif field in GLOBAL_FIELD_LISTS:
            result[field] = GLOBAL_FIELD_LISTS[field]
    return result


def export_workbook(
    database: str | Path,
    output: str | Path,
    *,
    actor: str = "codex",
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    database = Path(database).resolve()
    output = Path(output).resolve()
    conn = connect(database)
    try:
        filters = {
            key: value for key, value in (filters or {}).items()
            if value is not None and value != ""
        }
        complete = not filters
        baseline = catalog_state_hash(conn)
        schema = conn.execute("SELECT MAX(CAST(version AS INTEGER)) FROM schema_migrations").fetchone()[0]
        artifacts = _selected_artifacts(conn, filters)
        table_rows = _scoped_table_rows(conn, artifacts, complete=complete)
        wb = Workbook()
        wb.remove(wb.active)
        manifest_ws = wb.create_sheet(SHEETS[0])
        manifest_ws.append(["key", "value"])
        manifest = {
            "workbook_contract": "secureguide-catalog-workbook-v3",
            "schema_version": int(schema or 0),
            "baseline_db_sha256": baseline,
            "baseline_catalog_state_sha256": baseline,
            "database_name": database.name,
            "database_path": _portable_path(database),
            "exported_at_utc": datetime.now(timezone.utc).isoformat(),
            "exported_by": actor,
            "export_mode": "COMPLETE" if complete else "FILTERED",
            "export_scope": "ALL_CATALOG_ARTIFACTS" if complete else "DECLARED_FILTER_SCOPE",
            "filter_json": json.dumps(filters, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
            "quality_profile": filters.get("quality_profile") or "MINIMUM_VALID",
            "minimum_contract_sha256": file_hash(ROOT / "config" / "catalog_minimum_fields.yaml"),
            "source_manifest_sha256": file_hash(ROOT / "config" / "source_manifest.json"),
            "source_rights_sha256": file_hash(ROOT / "config" / "source_rights.yaml"),
            "artifact_count": len(table_rows["01_Artifacts"]),
            "canonical_total": conn.execute("SELECT COUNT(*) FROM security_artifacts").fetchone()[0],
            "raw_total": conn.execute("SELECT COUNT(*) FROM raw_artifacts").fetchone()[0],
            "disposition_total": conn.execute("SELECT COUNT(*) FROM raw_artifact_dispositions").fetchone()[0],
            "lineage_total": conn.execute("SELECT COUNT(*) FROM artifact_source_lineage").fetchone()[0],
            "scoped_disposition_count": len(table_rows["09_Raw_Dispositions"]),
            "scoped_lineage_count": len(table_rows["02_Source_Lineage"]),
            "excluded_data": "PROFILE_OPERATIONAL|RAW_PAYLOAD|DERIVED_EMBEDDINGS|WORKFLOW_BLUEPRINTS",
            "row_omission_semantics": "NO_CHANGE",
            "allowed_actions": "|".join(ACTION_VALUES),
        }
        manifest.update({
            f"row_count.{sheet}": len(rows)
            for sheet, rows in table_rows.items()
        })
        manifest["row_count.06_Type_Specific"] = len(table_rows["01_Artifacts"])
        for key, value in manifest.items():
            manifest_ws.append([key, value])

        for sheet in CORE_SHEETS[1:6]:
            table = TABLE_SHEETS[sheet]
            ws = wb.create_sheet(sheet)
            columns = _table_columns(conn, table)
            _append_table(ws, columns, table_rows[sheet])

        type_ws = wb.create_sheet("06_Type_Specific")
        type_rows = []
        artifact_ids = [row["id"] for row in table_rows["01_Artifacts"]]
        type_where = ""
        type_parameters: tuple[Any, ...] = ()
        if not complete:
            placeholders = ",".join("?" for _ in artifact_ids)
            type_where = (
                f" WHERE id IN ({placeholders})" if artifact_ids else " WHERE 0"
            )
            type_parameters = tuple(artifact_ids)
        for row in conn.execute(
            """SELECT id AS artifact_id,type,requirement_type,control_nature,
                      control_function,testability,asset_type,asset_criticality,
                      exception_approval_date,exception_expiry_date,effective_date
                 FROM security_artifacts"""
            + type_where
            + " ORDER BY id",
            type_parameters,
        ):
            type_rows.append(_row_dict(row, TYPE_FIELDS))
        _append_table(type_ws, list(TYPE_FIELDS), type_rows)

        refs = _reference_lists(conn)
        ref_ws = wb.create_sheet("07_Reference_Lists")
        ref_ws.append(["list_name", "code"])
        ranges: dict[str, tuple[int, int]] = {}
        for name in sorted(refs):
            start = ref_ws.max_row + 1
            for code in refs[name]:
                ref_ws.append([name, code])
            ranges[name] = (start, ref_ws.max_row)

        errors_ws = wb.create_sheet("08_Validation_Errors")
        errors_ws.append(["sheet", "row", "field", "code", "message"])

        disposition_ws = wb.create_sheet("09_Raw_Dispositions")
        _append_table(
            disposition_ws,
            _table_columns(conn, "raw_artifact_dispositions"),
            table_rows["09_Raw_Dispositions"],
        )

        for sheet, table in DETAIL_TABLE_SHEETS.items():
            ws = wb.create_sheet(sheet)
            _append_table(ws, _table_columns(conn, table), table_rows[sheet])

        for ws in wb.worksheets:
            _style_sheet(ws)
        action_fill = PatternFill("solid", fgColor="FFF2CC")
        for sheet in EDITABLE_SHEETS:
            ws = wb[sheet]
            for cell in ws["A"][1:]:
                cell.fill = action_fill
            actions = ACTION_VALUES if sheet == "01_Artifacts" else ACTION_VALUES[:2]
            action_dv = DataValidation(type="list", formula1=f'"{",".join(actions)}"')
            ws.add_data_validation(action_dv)
            action_dv.add(f"A2:A{max(ws.max_row + 500, 501)}")

        for name, (start, end) in ranges.items():
            safe = "REF_" + "".join(ch if ch.isalnum() else "_" for ch in name).upper()
            reference = f"{quote_sheetname('07_Reference_Lists')}!$B${start}:$B${end}"
            wb.defined_names.add(DefinedName(safe, attr_text=reference))
            for sheet in EDITABLE_SHEETS:
                ws = wb[sheet]
                headers = {cell.value: cell.column for cell in ws[1]}
                controlled = _field_lists_for_sheet(sheet, headers)
                matching_fields = [
                    field for field, list_name in controlled.items() if list_name == name
                ]
                for field in matching_fields:
                    col = get_column_letter(headers[field])
                    dv = DataValidation(type="list", formula1=f"={safe}", allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(f"{col}2:{col}{max(ws.max_row + 500, 501)}")

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return {
            "path": _portable_path(output), "sheetCount": len(wb.sheetnames),
            "sheets": wb.sheetnames, "baselineDbSha256": baseline,
            "workbookSha256": file_hash(output),
            "exportMode": manifest["export_mode"],
            "filters": filters,
        }
    finally:
        conn.close()


def _manifest(wb) -> dict[str, Any]:
    ws = wb["00_Manifest"]
    return {str(row[0].value): row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}


def _headers(ws) -> list[str]:
    return [str(cell.value) if cell.value is not None else "" for cell in ws[1]]


def _workbook_rows(ws) -> list[tuple[int, dict[str, Any]]]:
    headers = _headers(ws)
    rows: list[tuple[int, dict[str, Any]]] = []
    for number, cells in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(cell.value is None for cell in cells):
            continue
        rows.append((number, {headers[index]: _json_value(cell.value) for index, cell in enumerate(cells)}))
    return rows


def _business_values(row: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in row.items() if key not in META_COLUMNS}


def _key(sheet: str, row: dict[str, Any]) -> str:
    return canonical_hash({key: row.get(key) for key in PRIMARY_KEYS[sheet]})


def annotate_validation_errors(
    workbook: str | Path,
    errors: Iterable[dict[str, Any]],
    output: str | Path,
) -> dict[str, Any]:
    """Write an actionable, ordered error sheet to a workbook copy."""
    workbook = Path(workbook).resolve()
    output = Path(output).resolve()
    wb = load_workbook(workbook, data_only=False)
    if "08_Validation_Errors" not in wb.sheetnames:
        raise WorkbookError("08_Validation_Errors sheet is missing")
    ws = wb["08_Validation_Errors"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    sheet_order = {name: index for index, name in enumerate(SHEETS)}
    ordered = sorted(
        errors,
        key=lambda error: (
            sheet_order.get(str(error.get("sheet")), len(SHEETS)),
            int(error.get("row") or 0),
            str(error.get("field") or ""),
            str(error.get("code") or ""),
            str(error.get("message") or ""),
        ),
    )
    for error in ordered:
        ws.append(
            [
                error.get("sheet"),
                error.get("row"),
                error.get("field"),
                error.get("code"),
                error.get("message"),
            ]
        )
    _style_sheet(ws)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return {
        "path": _portable_path(output),
        "errorCount": len(ordered),
        "workbookSha256": file_hash(output),
    }


def validate_workbook(workbook: str | Path, database: str | Path) -> dict[str, Any]:
    workbook = Path(workbook).resolve()
    database = Path(database).resolve()
    wb = load_workbook(workbook, data_only=False)
    errors: list[dict[str, Any]] = []

    if wb.sheetnames != list(SHEETS):
        errors.append({"sheet": "WORKBOOK", "row": 0, "field": "sheetnames", "code": "SHEET_SET", "message": "Workbook sheets differ from the comprehensive contract order."})
        return {"valid": False, "errors": errors, "workbookSha256": file_hash(workbook)}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    errors.append({"sheet": ws.title, "row": cell.row, "field": cell.coordinate, "code": "FORMULA", "message": "Formulas are forbidden in curation workbooks."})

    manifest = _manifest(wb)
    if manifest.get("workbook_contract") != "secureguide-catalog-workbook-v3":
        errors.append({
            "sheet": "00_Manifest", "row": 2, "field": "workbook_contract",
            "code": "CONTRACT", "message": "Workbook contract must be secureguide-catalog-workbook-v3.",
        })
    conn = connect(database)
    try:
        current_db_hash = catalog_state_hash(conn)
        refs = _reference_lists(conn)
        table_columns = {sheet: _table_columns(conn, table) for sheet, table in TABLE_SHEETS.items()}
        table_columns["06_Type_Specific"] = list(TYPE_FIELDS)
        for sheet, columns in table_columns.items():
            expected_headers = [*META_COLUMNS, *columns]
            if _headers(wb[sheet]) != expected_headers:
                errors.append({"sheet": sheet, "row": 1, "field": "headers", "code": "HEADERS", "message": "Columns differ from the exported contract."})
            expected_count = manifest.get(f"row_count.{sheet}")
            actual_count = len(_workbook_rows(wb[sheet]))
            # Editable sheets intentionally permit row omission (NO_CHANGE) and
            # additions (UPSERT).  Their manifest counts describe the exported
            # baseline; export-only sheets must remain byte-for-byte complete.
            if expected_count is None or (
                sheet not in EDITABLE_SHEETS and int(expected_count) != actual_count
            ):
                errors.append({"sheet": sheet, "row": 0, "field": "row_count", "code": "ROW_COUNT", "message": "Manifest row count does not match the sheet."})
        if manifest.get("export_mode") == "COMPLETE":
            raw_total = conn.execute("SELECT COUNT(*) FROM raw_artifacts").fetchone()[0]
            disposition_count = len(_workbook_rows(wb["09_Raw_Dispositions"]))
            if disposition_count != raw_total:
                errors.append({"sheet": "09_Raw_Dispositions", "row": 0, "field": "raw_artifact_id", "code": "RAW_CLOSURE", "message": "Complete exports must contain exactly one disposition for every raw record."})
        elif manifest.get("export_mode") != "FILTERED":
            errors.append({"sheet": "00_Manifest", "row": 0, "field": "export_mode", "code": "EXPORT_MODE", "message": "Export mode must be COMPLETE or FILTERED."})
        for sheet in EDITABLE_SHEETS:
            ws = wb[sheet]
            expected_headers = [*META_COLUMNS, *table_columns[sheet]]
            if _headers(ws) != expected_headers:
                errors.append({"sheet": sheet, "row": 1, "field": "headers", "code": "HEADERS", "message": "Columns differ from the exported contract."})
                continue
            seen_keys: dict[str, int] = {}
            for row_number, row in _workbook_rows(ws):
                action = row.get("_action")
                if action not in ACTION_VALUES:
                    errors.append({"sheet": sheet, "row": row_number, "field": "_action", "code": "ACTION", "message": f"Invalid action {action}."})
                    continue
                if action == "DEPRECATE" and sheet != "01_Artifacts":
                    errors.append({"sheet": sheet, "row": row_number, "field": "_action", "code": "DEPRECATE_SCOPE", "message": "Only catalog artifacts support logical deprecation."})
                    continue
                values = _business_values(row)
                semantic_key = _key(sheet, values)
                if semantic_key in seen_keys:
                    errors.append({
                        "sheet": sheet,
                        "row": row_number,
                        "field": ",".join(PRIMARY_KEYS[sheet]),
                        "code": "DUPLICATE_ROW_KEY",
                        "message": (
                            "Editable row identity duplicates row "
                            f"{seen_keys[semantic_key]}."
                        ),
                    })
                else:
                    seen_keys[semantic_key] = row_number
                baseline_key = row.get("_baseline_key")
                if baseline_key and baseline_key != semantic_key:
                    errors.append({"sheet": sheet, "row": row_number, "field": PRIMARY_KEYS[sheet][0], "code": "IMMUTABLE_ID", "message": "Primary identity changed since export."})
                proposed_hash = row_hash(values)
                if action == "NO_CHANGE" and row.get("_baseline_hash") != proposed_hash:
                    errors.append({"sheet": sheet, "row": row_number, "field": "_action", "code": "ACTION_REQUIRED", "message": "Edited row must use UPSERT or DEPRECATE."})
                for field, list_name in _field_lists_for_sheet(sheet, values).items():
                    allowed = refs[list_name]
                    if values[field] not in (None, "") and values[field] not in allowed:
                        errors.append({"sheet": sheet, "row": row_number, "field": field, "code": "ENUM", "message": f"Value {values[field]} is not controlled by {list_name}."})
                if sheet == "01_Artifacts" and action == "UPSERT":
                    for field in load_contract()["core_required"]:
                        if values.get(field) in (None, ""):
                            errors.append({"sheet": sheet, "row": row_number, "field": field, "code": "MINIMUM", "message": "Required by the minimum catalog contract."})
                    confidence = values.get("classification_confidence")
                    if confidence is not None and float(confidence) <= 0.70 and not (
                        values.get("requires_human_review") in (1, True)
                        and values.get("ai_review_status") == "AIR-HUMAN-REVIEW"
                    ):
                        errors.append({"sheet": sheet, "row": row_number, "field": "classification_confidence", "code": "LOW_CONFIDENCE", "message": "Low confidence requires explicit human-review flags."})
    finally:
        conn.close()
    return {
        "valid": not errors, "errors": errors,
        "baselineDbSha256": manifest.get("baseline_db_sha256"),
        "currentDbSha256": current_db_hash,
        "workbookSha256": file_hash(workbook),
    }


def _current_row(conn: sqlite3.Connection, sheet: str, values: dict[str, Any]) -> dict[str, Any] | None:
    if sheet == "06_Type_Specific":
        row = conn.execute(
            "SELECT " + ",".join(f"{c if c != 'artifact_id' else 'id AS artifact_id'}" for c in TYPE_FIELDS)
            + " FROM security_artifacts WHERE id=?", (values.get("artifact_id"),)
        ).fetchone()
        return _row_dict(row, TYPE_FIELDS) if row else None
    table = TABLE_SHEETS[sheet]
    keys = PRIMARY_KEYS[sheet]
    if any(values.get(key) in (None, "") for key in keys):
        return None
    where = " AND ".join(f"{key}=?" for key in keys)
    row = conn.execute(
        f'SELECT * FROM "{table}" WHERE {where}', tuple(values[key] for key in keys)
    ).fetchone()
    return _row_dict(row) if row else None


def plan_workbook(
    workbook: str | Path, database: str | Path,
    resolutions: dict[str, str] | None = None,
) -> dict[str, Any]:
    validation = validate_workbook(workbook, database)
    resolutions = resolutions or {}
    if not validation["valid"]:
        raise WorkbookError(f"workbook validation failed with {len(validation['errors'])} error(s)")
    wb = load_workbook(workbook, data_only=False)
    conn = connect(database)
    try:
        entries: list[dict[str, Any]] = []
        conflicts: list[dict[str, Any]] = []
        db_stale = validation["baselineDbSha256"] != validation["currentDbSha256"]
        for sheet in EDITABLE_SHEETS:
            for row_number, row in _workbook_rows(wb[sheet]):
                action = row["_action"]
                if action == "NO_CHANGE":
                    continue
                values = _business_values(row)
                key = _key(sheet, values)
                current = _current_row(conn, sheet, values)
                current_hash = row_hash(current) if current else None
                resolution = resolutions.get(f"{sheet}:{key}") or resolutions.get("__database__")
                stale = db_stale or (
                    row.get("_baseline_hash") is not None
                    and current_hash != row.get("_baseline_hash")
                )
                if stale and resolution not in ("USE_WORKBOOK", "USE_DATABASE", "MANUAL"):
                    conflicts.append({"sheet": sheet, "row": row_number, "rowKey": key, "reason": "STALE_BASELINE"})
                entries.append({
                    "sheet": sheet, "row": row_number, "rowKey": key,
                    "action": action, "baselineHash": row.get("_baseline_hash"),
                    "currentHash": current_hash, "proposedHash": row_hash(values),
                    "values": values, "resolution": resolution,
                })
        plan = {
            "contract": "secureguide-catalog-workbook-plan-v3",
            "database": _portable_path(database),
            "workbook": _portable_path(workbook),
            "baselineDbSha256": validation["baselineDbSha256"],
            "currentDbSha256": validation["currentDbSha256"],
            "workbookSha256": validation["workbookSha256"],
            "entries": entries, "conflicts": conflicts,
        }
        plan["planSha256"] = canonical_hash(plan)
        return plan
    finally:
        conn.close()


def _update_artifact(conn: sqlite3.Connection, values: dict[str, Any]) -> None:
    artifact_id = values["id"]
    columns = _table_columns(conn, "security_artifacts")
    existing = conn.execute("SELECT 1 FROM security_artifacts WHERE id=?", (artifact_id,)).fetchone()
    writable = [c for c in columns if c in values and c != "id"]
    if existing:
        conn.execute(
            "UPDATE security_artifacts SET " + ",".join(f"{c}=?" for c in writable) + " WHERE id=?",
            tuple(values[c] for c in writable) + (artifact_id,),
        )
    else:
        insert_columns = [c for c in columns if c in values and values[c] is not None]
        conn.execute(
            f"INSERT INTO security_artifacts({','.join(insert_columns)}) VALUES({','.join('?' for _ in insert_columns)})",
            tuple(values[c] for c in insert_columns),
        )


def _apply_entry(conn: sqlite3.Connection, entry: dict[str, Any]) -> None:
    sheet, action, values = entry["sheet"], entry["action"], entry["values"]
    if entry.get("resolution") == "USE_DATABASE":
        return
    if action == "DEPRECATE":
        conn.execute(
            "UPDATE security_artifacts SET publication_status='DEPRECATED',is_active=0,"
            "updated_at=datetime('now'),version=version+1 WHERE id=?", (values["id"],)
        )
        return
    if sheet == "01_Artifacts":
        _update_artifact(conn, values)
    elif sheet == "06_Type_Specific":
        writable = [field for field in TYPE_FIELDS[1:] if field in values]
        conn.execute(
            "UPDATE security_artifacts SET " + ",".join(f"{f}=?" for f in writable) + " WHERE id=?",
            tuple(values[field] for field in writable) + (values["artifact_id"],),
        )
    else:
        table = TABLE_SHEETS[sheet]
        columns = [c for c in _table_columns(conn, table) if c in values and values[c] is not None]
        keys = PRIMARY_KEYS[sheet]
        current = _current_row(conn, sheet, values)
        if current:
            writable = [c for c in columns if c not in keys]
            if writable:
                conn.execute(
                    f"UPDATE {table} SET " + ",".join(f"{c}=?" for c in writable)
                    + " WHERE " + " AND ".join(f"{k}=?" for k in keys),
                    tuple(values[c] for c in writable) + tuple(values[k] for k in keys),
                )
        else:
            conn.execute(
                f'INSERT INTO "{table}"({",".join(columns)}) VALUES({",".join("?" for _ in columns)})',
                tuple(values[c] for c in columns),
            )


def _affected_artifact_ids(sheet: str, values: dict[str, Any]) -> set[str]:
    if sheet == "01_Artifacts":
        return {str(values["id"])} if values.get("id") not in (None, "") else set()
    if sheet == "04_Relationships":
        return {
            str(values[field])
            for field in ("source_id", "target_id")
            if values.get(field) not in (None, "")
        }
    artifact_id = values.get("artifact_id")
    return {str(artifact_id)} if artifact_id not in (None, "") else set()


def apply_workbook_plan(plan: dict[str, Any], *, actor: str = "codex") -> dict[str, Any]:
    database = Path(plan["database"]).resolve()
    if database == RELEASE_ASSET:
        raise WorkbookError("direct curation of mobile/assets/catalog.db is forbidden")
    if plan.get("contract") != "secureguide-catalog-workbook-plan-v3":
        raise WorkbookError("unsupported workbook plan contract")
    if canonical_hash({k: v for k, v in plan.items() if k != "planSha256"}) != plan.get("planSha256"):
        raise WorkbookError("plan hash mismatch")
    unresolved = [c for c in plan["conflicts"] if not any(
        e["rowKey"] == c["rowKey"] and e.get("resolution") in ("USE_WORKBOOK", "USE_DATABASE", "MANUAL")
        for e in plan["entries"]
    )]
    if unresolved:
        raise WorkbookConflict(f"{len(unresolved)} unresolved workbook conflict(s)")
    if file_hash(plan["workbook"]) != plan["workbookSha256"]:
        raise WorkbookConflict("workbook changed after plan creation")

    conn = connect(database)
    run_id = f"WBR-{uuid.uuid4()}"
    try:
        conn.execute("BEGIN IMMEDIATE")
        current_state = catalog_state_hash(conn)
        if current_state != plan["currentDbSha256"]:
            raise WorkbookConflict("database changed after plan creation")
        conn.execute(
            """INSERT INTO catalog_workbook_runs(
                   id,operation,workbook_path,baseline_db_sha256,workbook_sha256,
                   status,actor,conflict_resolution_json
               ) VALUES(?,?,?,?,?,'STARTED',?,?)""",
            (run_id, "APPLY", plan["workbook"], plan["baselineDbSha256"],
             plan["workbookSha256"], actor,
             json.dumps({e["rowKey"]: e.get("resolution") for e in plan["entries"] if e.get("resolution")}, sort_keys=True)),
        )
        affected: set[str] = set()
        for entry in plan["entries"]:
            if entry.get("resolution") == "MANUAL":
                raise WorkbookConflict(f"manual resolution incomplete for {entry['rowKey']}")
            _apply_entry(conn, entry)
            affected.update(_affected_artifact_ids(entry["sheet"], entry["values"]))
            conn.execute(
                """INSERT INTO catalog_workbook_row_audit(
                       run_id,sheet_name,row_key,action,baseline_hash,current_hash,
                       proposed_hash,outcome,resolution
                   ) VALUES(?,?,?,?,?,?,?,'APPLIED',?)""",
                (run_id, entry["sheet"], entry["rowKey"], entry["action"],
                 entry["baselineHash"], entry["currentHash"], entry["proposedHash"],
                 entry.get("resolution")),
            )
        contract = load_contract()
        invalid: list[dict[str, Any]] = []
        for artifact_id in sorted(affected):
            row = conn.execute("SELECT * FROM security_artifacts WHERE id=?", (artifact_id,)).fetchone()
            if row and row["is_active"]:
                result = minimum_result(conn, row, contract)
                if not result["valid"]:
                    invalid.append({"id": artifact_id, **result})
        if invalid:
            raise WorkbookError(f"post-apply minimum validation failed: {invalid[:3]}")
        missing_dispositions = conn.execute(
            """SELECT COUNT(*) FROM raw_artifacts r WHERE NOT EXISTS(
                   SELECT 1 FROM raw_artifact_dispositions d
                    WHERE d.raw_artifact_id=r.id)"""
        ).fetchone()[0]
        inconsistent_lineage = conn.execute(
            """SELECT COUNT(*) FROM raw_artifact_dispositions d
                 WHERE d.disposition IN ('SUPPORTS_CANONICAL','SPLIT') AND NOT EXISTS(
                   SELECT 1 FROM artifact_source_lineage l
                    WHERE l.raw_artifact_id=d.raw_artifact_id
                      AND l.lineage_role=d.disposition)"""
        ).fetchone()[0]
        if missing_dispositions or inconsistent_lineage:
            raise WorkbookError(
                "post-apply raw disposition closure failed: "
                f"missing={missing_dispositions}, inconsistent={inconsistent_lineage}"
            )
        summary = {"entries": len(plan["entries"]), "affectedArtifacts": len(affected)}
        conn.execute(
            "UPDATE catalog_workbook_runs SET status='APPLIED',summary_json=?,completed_at=datetime('now') WHERE id=?",
            (json.dumps(summary, sort_keys=True), run_id),
        )
        conn.execute("COMMIT")
        return {"status": "APPLIED", "runId": run_id, **summary, "catalogStateSha256": catalog_state_hash(conn)}
    except Exception:
        if conn.in_transaction:
            conn.execute("ROLLBACK")
        raise
    finally:
        conn.close()
